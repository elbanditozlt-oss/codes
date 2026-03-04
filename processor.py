# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from pathlib import Path
import numpy as np
import re
import calendar
from datetime import date, datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# External fetch for Google Sheets (optional)
author_has_requests = True
try:
    import requests  # type: ignore
except Exception:
    author_has_requests = False

# ===============================
# НАСТРОЙКИ
# ===============================
APP_VERSION = "v3.1.0"
MAIN_SHEET_NAME = "Сводная мотивация"
OKLAD_SHEET_NAME = "Оклады"
LOCAL_OKLADS_FILE = Path("оклады.xlsx")
USE_VLOOKUP_FALLBACK_DEFAULT = False

# Переработки (Google Sheets)
OVERTIME_SHEET_NAME = "Переработки инженеры"
# Жёстко удаляемые колонки из ДЕТАЛЬНОЙ таблицы при записи листа
OVERTIME_DROP_ALWAYS = ["US", "KKT", "USKKT"]
# «Мусор» для удаления ещё на этапе объединения (если встретится)
OVERTIME_DROP_COLS = [
    "Кол. вып. заявок в р.время",
    "Километраж",
    "Текущий статус",
    "Кто согласовал",
    "Дата согласования",
    "Заработок инженера",
    "Рекомендовано часов",
    "Согласовано часов (Всего)",
    "Комментарий руководителя",
]
OVERTIME_POS_CANDIDATES = ["POS", "Количество часов", "Часы", "Кол-во часов", "Итого часов", "Переработки (часы)"]
OVERTIME_FIO_COL = "ФИО"

# ===============================
# УТИЛИТЫ
# ===============================
def normalize_fio(s: str):
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return s
    return " ".join(str(s).split()).strip()

def read_main_workbook(uploaded_file):
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
        if MAIN_SHEET_NAME not in wb.sheetnames:
            st.error("Нет листа «Сводная мотивация».")
            return None
        return wb
    except Exception as e:
        st.error(f"Ошибка загрузки Excel: {e}")
        return None

def ws_header_map(ws):
    hmap = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None and str(v).strip() != "":
            name = str(v).strip()
            if name not in hmap:
                hmap[name] = col
    return hmap

def get_last_row_by_fio(ws, fio_col_idx=2):
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(row=r, column=fio_col_idx).value
        if v not in (None, ""):
            return r
    return 1

def clean_fio(ws, fio_col_idx, last_row):
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=fio_col_idx).value = normalize_fio(
            ws.cell(row=r, column=fio_col_idx).value
        )

def delete_rows_partial(ws, fio_col_idx, patterns):
    filters = []
    for p in patterns:
        parts = [x.strip().lower() for x in p.split() if x.strip()]
        if parts:
            filters.append(parts)
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=fio_col_idx).value
        if not val:
            continue
        fio_norm = normalize_fio(val).lower()
        for parts in filters:
            if all(part in fio_norm for part in parts):
                rows_to_delete.append(r)
                break
    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)
    return len(rows_to_delete)

def set_zero(ws, col_idx_list, last_row):
    changed = 0
    for col in col_idx_list:
        if col is None:
            continue
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col).value = 0
            changed += 1
    return changed

def read_local_oklads():
    if not LOCAL_OKLADS_FILE.exists():
        st.error("❗ Файл «оклады.xlsx» не найден рядом с приложением.")
        return None
    try:
        return pd.read_excel(LOCAL_OKLADS_FILE, engine="openpyxl")
    except Exception as e:
        st.error(f"Ошибка чтения «оклады.xlsx»: {e}")
        return None

def ensure_oklad_sheet(df_ok):
    if df_ok is None or df_ok.empty:
        return df_ok
    cols = df_ok.columns.tolist()
    fio = "ФИО" if "ФИО" in cols else cols[0]
    sm  = "Сумма" if "Сумма" in cols else (cols[1] if len(cols) > 1 else cols[0])
    return df_ok[[fio, sm] + [c for c in cols if c not in (fio, sm)]].copy()

def add_oklad_sheet(wb, df_ok):
    if OKLAD_SHEET_NAME in wb.sheetnames:
        del wb[OKLAD_SHEET_NAME]
    ws = wb.create_sheet(OKLAD_SHEET_NAME)
    for c_idx, col in enumerate(df_ok.columns, 1):
        ws.cell(row=1, column=c_idx).value = col
    for r_idx, (_, row) in enumerate(df_ok.iterrows(), 2):
        for c_idx, col in enumerate(df_ok.columns, 1):
            ws.cell(row=r_idx, column=c_idx).value = row[col]
    # скрываем лист «Оклады»
    ws.sheet_state = "hidden"

def remove_sheet_protection(ws):
    try:
        ws.protection.sheet = False
        ws.protection.enable = False
        ws.protection.password = None
    except Exception:
        pass

def finalize_sheets(wb, df_ok_sheet):
    deleted = []
    try:
        for name in list(wb.sheetnames):
            if name != MAIN_SHEET_NAME:
                deleted.append(name)
                del wb[name]
    except Exception as e:
        st.error("❗ Не удалось удалить лишние листы. Возможно, включена защита структуры книги.\n"
                 f"Техническая деталь: {e}")
        return False, deleted, False
    added_oklads = False
    try:
        if df_ok_sheet is not None and not df_ok_sheet.empty:
            add_oklad_sheet(wb, df_ok_sheet)
            added_oklads = True
        return True, deleted, added_oklads
    except Exception as e:
        st.error("❗ Не удалось добавить лист «Оклады». Возможно, структура книги под защитой.\n"
                 f"Техническая деталь: {e}")
        return False, deleted, added_oklads

def preview_df(ws, last_row):
    rows = []
    for r in range(1, last_row + 1):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])
    if not rows:
        return pd.DataFrame()
    header = rows[0]
    data = rows[1:]
    cols = []
    seen = {}
    for h in header:
        nm = "Unnamed" if h in (None, "", "None") else str(h).strip()
        if nm in seen:
            seen[nm] += 1
            nm = f"{nm}_{seen[nm]}"
        else:
            seen[nm] = 1
        cols.append(nm)
    try:
        return pd.DataFrame(data, columns=cols)
    except Exception:
        return pd.DataFrame(data)

def build_formula_xlookup(i: int) -> str:
    return (
        f"=IFERROR(XLOOKUP($B{i},'{OKLAD_SHEET_NAME}'!$A:$A,'{OKLAD_SHEET_NAME}'!$B:$B,\"\"),\"\")"
    )

def build_formula_vlookup(i: int) -> str:
    return (
        f"=IFERROR(VLOOKUP($B{i},'{OKLAD_SHEET_NAME}'!$A:$B,2,FALSE),\"\")"
    )

def append_suffix_to_filename(original_name: str, suffix: str) -> str:
    if "." not in original_name:
        return original_name + suffix
    base, ext = original_name.rsplit(".", 1)
    return f"{base}{suffix}.{ext}"

# ---------- Форматирование часов ----------

def _round_to_half(x: float) -> float:
    if pd.isna(x):
        return 0.0
    return float(np.round(x * 2) / 2.0)

def _format_half_ru(x: float) -> str:
    if pd.isna(x):
        return ""
    xi = int(round(x))
    if abs(x - xi) < 1e-9:
        return str(xi)
    xh = _round_to_half(x)
    return str(xh).replace(".", ",")

# -------------------- Переработки: загрузка и обработка --------------------

def _extract_gsheet_file_id(url: str) -> str | None:
    if not url:
        return None
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    return m.group(1) if m else None

def _make_gsheet_export_xlsx_url(url: str) -> str | None:
    if "export?format=xlsx" in url:
        return url
    file_id = _extract_gsheet_file_id(url)
    if not file_id:
        return None
    return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"

def fetch_gsheet_as_xlsx_bytes(url: str) -> BytesIO:
    if not author_has_requests:
        raise RuntimeError("В окружении отсутствует модуль requests — установите его (pip install requests)")
    export_url = _make_gsheet_export_xlsx_url(url)
    if not export_url:
        raise ValueError("Не удалось распознать ссылку Google Sheets.")
    resp = requests.get(export_url, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"Google Sheets вернул статус {resp.status_code}")
    return BytesIO(resp.content)

def _find_first_present_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    # точное совпадение
    cols_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.strip().lower()
        if key in cols_map:
            return cols_map[key]
    # мягкое вхождение
    low_cols = [(c, str(c).strip().lower()) for c in df.columns]
    for cand in candidates:
        key = cand.strip().lower()
        for orig, low in low_cols:
            if key in low:
                return orig
    return None

def _coerce_pos(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace(",", ".", regex=False).str.strip()
    s = pd.to_numeric(s, errors="coerce").fillna(0.0)
    return s

def load_and_prepare_overtime_from_gsheet(url: str) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    xlsx_bytes = fetch_gsheet_as_xlsx_bytes(url)
    xls = pd.ExcelFile(xlsx_bytes, engine="openpyxl")
    sheet_names = xls.sheet_names
    df_list = []
    for name in sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=name, engine="openpyxl")
            if df is None or df.empty:
                continue
            df_list.append(df)
        except Exception:
            continue
    if not df_list:
        raise RuntimeError("Не удалось прочитать ни одной вкладки из Google Sheets.")

    df_all = pd.concat(df_list, ignore_index=True)
    rows_before = len(df_all)

    # убираем лишние колонки ещё на этапе объединения
    drop_cols_exist = [c for c in OVERTIME_DROP_COLS if c in df_all.columns]
    if drop_cols_exist:
        df_all = df_all.drop(columns=drop_cols_exist, errors="ignore")

    fio_col = _find_first_present_column(df_all, [OVERTIME_FIO_COL])
    if fio_col is None:
        raise RuntimeError("В таблице переработок не найдена колонка «ФИО».")

    pos_col = _find_first_present_column(df_all, OVERTIME_POS_CANDIDATES)
    if pos_col is None:
        raise RuntimeError("В таблице переработок не найдена колонка часов (POS/Количество часов/Итого часов).")

    df_all[fio_col] = df_all[fio_col].astype(str)
    df_all = df_all[df_all[fio_col].str.strip() != ""].copy()

    df_all[pos_col] = _coerce_pos(df_all[pos_col])
    rows_after = len(df_all)

    df_detail = df_all.copy()

    df_summary = (
        df_all.groupby(fio_col, as_index=False)[pos_col]
        .sum()
        .rename(columns={fio_col: "ФИО", pos_col: "Переработки (часы)"})
        .sort_values("Переработки (часы)", ascending=False, kind="mergesort")
    )

    # Округляем и формируем «красивую» колонку
    df_summary["Переработки (часы)"] = df_summary["Переработки (часы)"].apply(_round_to_half)
    df_summary["Переработки (часы) — вывод"] = df_summary["Переработки (часы)"].apply(_format_half_ru)

    metrics = {
        "sheets": len(sheet_names),
        "rows_before": rows_before,
        "rows_after": rows_after,
        "dropped_cols": drop_cols_exist,
        "fio_col": fio_col,
        "pos_col": pos_col,
        "unique_fio": df_summary["ФИО"].nunique(),
        "sum_pos": float(df_summary["Переработки (часы)"].sum()),
    }
    return df_detail, df_summary, metrics

# ---------- Оформление листа «Переработки инженеры» ----------

def add_overtime_sheet(wb, df_summary: pd.DataFrame, df_detail: pd.DataFrame, sheet_name: str = OVERTIME_SHEET_NAME):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # СВОДНАЯ
    ws.cell(row=1, column=1, value="ФИО").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Переработки (часы)").font = Font(bold=True)

    for i, row in enumerate(df_summary.itertuples(index=False), start=2):
        ws.cell(row=i, column=1, value=row[0])
        txt = row[2] if len(row) >= 3 else _format_half_ru(row[1])
        ws.cell(row=i, column=2, value=txt)

    # Итоговая строка
    total_numeric = float(df_summary["Переработки (часы)"].sum()) if "Переработки (часы)" in df_summary.columns else 0.0
    total_text = _format_half_ru(total_numeric)
    total_row = 2 + len(df_summary)
    ws.cell(row=total_row, column=1, value="Общий итог").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=total_text).font = Font(bold=True)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    for r in range(1, total_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")

    # Рамки сводной: тонкие внутри + толстая внешняя; шапка с заливкой
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    s_top, s_left, s_bottom, s_right = 1, 1, total_row, 2
    for r in range(s_top, s_bottom + 1):
        for c in range(s_left, s_right + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if r == s_top:
                cell.border = Border(left=cell.border.left, right=cell.border.right, top=thick, bottom=cell.border.bottom)
            if r == s_bottom:
                cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=thick)
            if c == s_left:
                cell.border = Border(left=thick, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
            if c == s_right:
                cell.border = Border(left=cell.border.left, right=thick, top=cell.border.top, bottom=cell.border.bottom)

    for c in range(s_left, s_right + 1):
        ws.cell(row=1, column=c).fill = header_fill

    # ДЕТАЛЬНАЯ: удаляем жёстко заданные колонки
    detail_df = df_detail.copy()
    drop_always = [c for c in OVERTIME_DROP_ALWAYS if c in detail_df.columns]
    if drop_always:
        detail_df = detail_df.drop(columns=drop_always, errors="ignore")

    if detail_df is not None and not detail_df.empty:
        detail_header_row = total_row + 2
        # Заголовки
        for j, col in enumerate(detail_df.columns, start=1):
            cell = ws.cell(row=detail_header_row, column=j, value=str(col))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = header_fill
        # Данные
        for r_idx, row in enumerate(detail_df.itertuples(index=False), start=detail_header_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        # Автофильтр
        last_row = detail_header_row + 1 + len(detail_df)
        last_col_letter = get_column_letter(len(detail_df.columns))
        ws.auto_filter.ref = f"A{detail_header_row}:{last_col_letter}{last_row}"
        # Рамки тонкие
        d_top, d_left, d_bottom, d_right = detail_header_row, 1, last_row, len(detail_df.columns)
        for r in range(d_top, d_bottom + 1):
            for c in range(d_left, d_right + 1):
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Ширина столбцов
        for c in range(1, len(detail_df.columns) + 1):
            col_letter = get_column_letter(c)
            if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < 16:
                ws.column_dimensions[col_letter].width = 18

# ---------- Лист «Лог» ----------

def add_log_sheet(wb, changes_log, meta: dict):
    name = "Лог"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    ws["A1"].value = "Отчёт об обработке файла"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    rows_meta = [
        ("Дата и время", meta.get("timestamp", "")),
        ("Версия приложения", meta.get("app_version", "")),
        ("Исходный файл", meta.get("filename", "")),
        ("Месяц/Год", f"{meta.get('month_name', '')} {meta.get('year', '')}"),
        ("Рабочие дни (Пн–Пт, минус праздники РФ)", meta.get("workdays", "")),
        ("Режим формул", "VLOOKUP" if meta.get("use_vlookup") else "XLOOKUP"),
    ]
    start_row = 3
    for i, (k, v) in enumerate(rows_meta, start=start_row):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=2).value = v
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="left")

    body_row = start_row + len(rows_meta) + 1
    ws.cell(row=body_row, column=1).value = "Изменения"
    ws.cell(row=body_row, column=1).font = Font(bold=True)
    for j, line in enumerate(changes_log, start=body_row + 1):
        ws.cell(row=j, column=1).value = f"• {line}"

    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22

    # скрываем лист «Лог»
    ws.sheet_state = "hidden"


def ensure_sheet_order(wb, desired_order: list[str]):
    present = {name: wb[name] for name in wb.sheetnames}
    ordered = [present[name] for name in desired_order if name in present]
    for nm in wb.sheetnames:
        if nm not in desired_order:
            ordered.append(present[nm])
    wb._sheets = ordered  # type: ignore

# ===============================
# UI
# ===============================
st.set_page_config(page_title=f"Обработчик мотивации {APP_VERSION}", layout="wide")
st.title(f"📊 Обработчик мотивации — {APP_VERSION}")

with st.sidebar:
    st.header("⚙️ Настройки")
    default_bad = (
        "Хамидуллин Руслан\n"
        "Селюх Артем\n"
        "Орлов Дмитрий\n"
        "Чиканов Алексей\n"
        "Захаров Никита"
    )
    bad_text = st.text_area("ФИО для удаления (каждое с новой строки):", default_bad, height=150)
    remove_bad = st.checkbox("Удалять строки по списку ФИО", value=False)
    use_vlookup_ui = st.checkbox("VLOOKUP вместо XLOOKUP", value=USE_VLOOKUP_FALLBACK_DEFAULT)

uploaded_file = st.file_uploader("Загрузите файл Excel (.xlsx) с листом «Сводная мотивация»", type=["xlsx"])

if "overtime" not in st.session_state:
    st.session_state["overtime"] = {"ready": False, "detail": None, "summary": None, "metrics": None}

# ======================================================================
# ОСНОВНАЯ ОБРАБОТКА
# ======================================================================
if uploaded_file:
    filename_lower = uploaded_file.name.lower()
    original_filename = uploaded_file.name

    months = {
        "январ": 1, "феврал": 2, "март": 3, "апрел": 4, "май": 5, "мая": 5,
        "июн": 6, "июл": 7, "август": 8, "сентябр": 9, "октябр": 10,
        "ноябр": 11, "декабр": 12
    }
    month = None
    for key, value in months.items():
        if key in filename_lower:
            month = value
            break

    year_match = re.search(r"20\d{2}", filename_lower)
    year = int(year_match.group(0)) if year_match else date.today().year

    if not month:
        st.error("❗ Не удалось определить месяц из названия файла. Укажите месяц в имени файла (например, «март 2026»).")
        st.stop()

    holidays = []
    if month == 1:
        holidays += [date(year, 1, d) for d in range(1, 8)]
    if month == 2:
        holidays.append(date(year, 2, 23))
    if month == 3:
        holidays.append(date(year, 3, 8))
    if month == 5:
        holidays += [date(year, 5, 1), date(year, 5, 9)]
    if month == 6:
        holidays.append(date(year, 6, 12))
    if month == 11:
        holidays.append(date(year, 11, 4))

    _, days_in_month = calendar.monthrange(year, month)
    workdays = sum(
        1
        for d in range(1, days_in_month + 1)
        if (date(year, month, d).weekday() < 5) and (date(year, month, d) not in holidays)
    )

    month_name_ru = {
        1: "январь", 2: "февраль", 3: "март", 4: "апрель", 5: "май",
        6: "июнь", 7: "июль", 8: "август", 9: "сентябрь", 10: "октябрь",
        11: "ноябрь", 12: "декабрь"
    }.get(month, "")

    wb = read_main_workbook(uploaded_file)
    if wb is None:
        st.stop()

    ws = wb[MAIN_SHEET_NAME]
    remove_sheet_protection(ws)
    hmap = ws_header_map(ws)

    FIO_COL = 2
    AC_COL  = hmap.get("Заработная плата инженера", 29)

    last_row = get_last_row_by_fio(ws, FIO_COL)

    st.subheader("Предпросмотр (до изменений)")
    st.dataframe(preview_df(ws, last_row), use_container_width=True)

    df_ok = read_local_oklads()
    if df_ok is None:
        st.stop()
    df_ok_sheet = ensure_oklad_sheet(df_ok)

    st.subheader("Лист «Оклады» (будет создан/обновлён)")
    st.dataframe(df_ok_sheet, use_container_width=True)

    st.subheader("Переработки (опционально)")
    use_overtime = st.radio("Загрузить переработки из Google Sheets?", options=["Нет", "Да"], horizontal=True, index=0)
    if use_overtime == "Да":
        gsheet_url = st.text_input("Вставьте ссылку на Google Sheets (доступ по ссылке):", value="", help="Ссылка вида https://docs.google.com/spreadsheets/d/…")
        if st.button("Загрузить переработки"):
            try:
                df_detail, df_summary, metrics = load_and_prepare_overtime_from_gsheet(gsheet_url)
                st.session_state["overtime"] = {
                    "ready": True,
                    "detail": df_detail,
                    "summary": df_summary,
                    "metrics": metrics,
                }
                st.success(
                    f"Переработки загружены: вкладок {metrics['sheets']}, строк до очистки {metrics['rows_before']}, после {metrics['rows_after']}, уникальных ФИО {metrics['unique_fio']}, суммарно часов {metrics['sum_pos']:.2f}."
                )
                st.info(f"Опознаны колонки → ФИО: {metrics.get('fio_col')!r}, Часы: {metrics.get('pos_col')!r}")
                with st.expander("Предпросмотр сводной (ТОП‑20):", expanded=True):
                    cols = ["ФИО", "Переработки (часы) — вывод"]
                    st.dataframe(
                        df_summary[cols].rename(columns={"Переработки (часы) — вывод": "Переработки (часы)"}).head(20),
                        use_container_width=True,
                    )
                with st.expander("Предпросмотр детальной таблицы (первые 50 строк):", expanded=False):
                    st.dataframe(df_detail.head(50), use_container_width=True)
                if metrics.get("dropped_cols"):
                    st.info("Удалены колонки при объединении: " + ", ".join(metrics["dropped_cols"]))
            except Exception as e:
                st.session_state["overtime"] = {"ready": False, "detail": None, "summary": None, "metrics": None}
                st.error(f"Не удалось загрузить переработки: {e}")
    else:
        st.session_state["overtime"] = {"ready": False, "detail": None, "summary": None, "metrics": None}

    if st.button("Сформировать processed.xlsx"):
        changes_log = []
        meta = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "app_version": APP_VERSION,
            "filename": original_filename,
            "month_name": month_name_ru,
            "year": year,
            "workdays": workdays,
            "use_vlookup": use_vlookup_ui,
        }

        try:
            clean_fio(ws, FIO_COL, last_row)
            processed_rows = max(last_row - 1, 0)
            changes_log.append(f"Нормализованы ФИО (обработано строк: {processed_rows}).")
        except Exception as e:
            st.error(f"Ошибка при нормализации ФИО: {e}")
            changes_log.append(f"❗ Ошибка при нормализации ФИО: {e}")

        try:
            if remove_bad:
                patterns = [x.strip() for x in bad_text.splitlines() if x.strip()]
                removed = delete_rows_partial(ws, FIO_COL, patterns)
                changes_log.append(f"Удалено строк по списку ФИО: {removed}.")
                last_row = get_last_row_by_fio(ws, FIO_COL)
            else:
                changes_log.append("Удаление строк по списку ФИО пропущено (выключено).")
        except Exception as e:
            st.error(f"Ошибка при удалении строк по ФИО: {e}")
            changes_log.append(f"❗ Ошибка при удалении строк по ФИО: {e}")

        try:
            zeroed_cells = set_zero(ws, [20, 21], last_row)
            changes_log.append(f"Обнулены столбцы T и U (установлено нулей: {zeroed_cells}).")
        except Exception as e:
            st.error(f"Ошибка при обнулении T/U: {e}")
            changes_log.append(f"❗ Ошибка при обнулении T/U: {e}")

        V_COL = 22
        try:
            days_filled = 0
            for r in range(2, last_row + 1):
                v = ws.cell(row=r, column=V_COL).value
                if v in (0, None, "", "0"):
                    ws.cell(row=r, column=V_COL).value = workdays
                    days_filled += 1
            changes_log.append(f"Подставлены рабочие дни ({workdays}) в {days_filled} строках.")
        except Exception as e:
            st.error(f"Ошибка при заполнении рабочих дней в V: {e}")
            changes_log.append(f"❗ Ошибка при заполнении рабочих дней в V: {e}")

        try:
            if AC_COL is None:
                AC_COL = 29
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=AC_COL).value = None
            use_vlookup = use_vlookup_ui
            formulas_written = 0
            for r in range(2, last_row + 1):
                formula = build_formula_vlookup(r) if use_vlookup else build_formula_xlookup(r)
                ws.cell(row=r, column=AC_COL).value = formula
                formulas_written += 1
            changes_log.append(
                f"Перезаписаны формулы в столбце AC ({'VLOOKUP' if use_vlookup else 'XLOOKUP'}), строк: {formulas_written}."
            )
        except Exception as e:
            st.error(f"Ошибка при вставке формул в AC: {e}")
            changes_log.append(f"❗ Ошибка при вставке формул в AC: {e}")

        try:
            R_COL = 18
            YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            painted = 0
            for r in range(2, last_row + 1):
                cell = ws.cell(row=r, column=R_COL)
                val = getattr(cell, "internal_value", None)
                if val is None:
                    val = cell.value
                num = None
                try:
                    if isinstance(val, str):
                        num = float(val.replace(",", ".").strip())
                    else:
                        num = float(val)
                except Exception:
                    num = None
                if num is not None and num < 8:
                    cell.fill = YELLOW
                    painted += 1
            changes_log.append(f"Окрашено ячеек в столбце R (значение < 8.00): {painted}.")
        except Exception as e:
            st.error(f"Ошибка при окрашивании R: {e}")
            changes_log.append(f"❗ Ошибка при окрашивании R: {e}")

        ok_finalize = False
        deleted_sheets = []
        oklads_added = False
        try:
            ok_finalize, deleted_sheets, oklads_added = finalize_sheets(wb, df_ok_sheet)
            if ok_finalize:
                if deleted_sheets:
                    changes_log.append(f"Удалены листы: {', '.join(deleted_sheets)}.")
                else:
                    changes_log.append("Лишние листы для удаления не найдены.")
                changes_log.append("Добавлен лист «Оклады»." if oklads_added else "Лист «Оклады» не добавлен (нет данных).")
            else:
                changes_log.append("❗ Не удалось финализировать структуру книги (см. ошибки выше).")
        except Exception as e:
            st.error(f"Ошибка при финализации структуры книги: {e}")
            changes_log.append(f"❗ Ошибка при финализации структуры книги: {e}")

        try:
            overtime_state = st.session_state.get("overtime", {})
            if overtime_state.get("ready") and overtime_state.get("summary") is not None and overtime_state.get("detail") is not None:
                add_overtime_sheet(wb, overtime_state["summary"], overtime_state["detail"], sheet_name=OVERTIME_SHEET_NAME)
                changes_log.append("Добавлен лист «Переработки инженеры».")
            else:
                changes_log.append("Переработки не были загружены — лист «Переработки инженеры» не добавлен.")
        except Exception as e:
            st.warning(f"Лист «{OVERTIME_SHEET_NAME}» не добавлен (ошибка): {e}")
            changes_log.append(f"❗ Ошибка добавления листа «{OVERTIME_SHEET_NAME}»: {e}")

        try:
            ws = wb[MAIN_SHEET_NAME]
            ws.auto_filter.ref = "A1:AM1"
            changes_log.append("Установлен автофильтр диапазона A1:AM1 на листе «Сводная мотивация».")
        except Exception as e:
            changes_log.append(f"Предупреждение: автофильтр не установлен (техническая деталь: {e}).")

        try:
            add_log_sheet(wb, changes_log, meta)
            changes_log.append("Создан лист «Лог» с отчётом об обработке (скрыт).")
        except Exception as e:
            st.error(f"Ошибка при создании листа «Лог»: {e}")
            changes_log.append(f"❗ Ошибка при создании листа «Лог»: {e}")

        try:
            ensure_sheet_order(wb, [MAIN_SHEET_NAME, OKLAD_SHEET_NAME, OVERTIME_SHEET_NAME, "Лог"])
            changes_log.append("Установлен порядок листов: Сводная мотивация → Оклады → Переработки инженеры → Лог.")
        except Exception:
            pass

        try:
            out = BytesIO()
            wb.save(out)
            st.success("Готово! Переработки обработаны, сводная с итогом оформлена, лишние колонки удалены, «Оклады» и «Лог» скрыты.")
            download_name = append_suffix_to_filename(original_filename, " — processed")
            st.download_button(
                label=f"Скачать {download_name}",
                data=out.getvalue(),
                file_name=download_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Ошибка сохранения файла: {e}")
            changes_log.append(f"❗ Ошибка сохранения файла: {e}")

        st.subheader("Отчёт о внесённых изменениях")
        for line in changes_log:
            st.write("• " + line)
